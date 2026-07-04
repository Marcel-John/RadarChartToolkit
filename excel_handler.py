from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from excel_validator import validate_chart_excel

from models import (
    RadarChart,
    RadarData,
    RadarDataset,
    RadarStyle,
)


def _load_meta(sheet: Worksheet) -> str:
    """Liest den Titel aus dem Meta-Sheet (Zelle B2)."""
    title = sheet["B2"].value
    
    if title is None:
        raise ValueError("Missing chart title.")
    
    return str(title)


def _load_data(sheet: Worksheet) -> tuple[list[str], list[RadarDataset]]:
    """Liest Labels und Datensätze (inkl. Farbe, Marker und Linienstärke) aus dem Data-Sheet."""
    labels = []

    # Zeile 1: Namen
    dataset_names = [str(cell.value) for cell in sheet[1][1:]]
    
    # Zeile 2-4: Optische Eigenschaften (mit Fallback-Werten und Typ-Absicherung für den Linter)
    dataset_colors = [str(cell.value) if cell.value is not None else "#007ACC" for cell in sheet[2][1:]]
    dataset_markers = [str(cell.value) if cell.value is not None else "o" for cell in sheet[3][1:]]
    
    # Das isinstance() löst die ConvertibleToFloat-Warnung deines Typprüfers auf
    dataset_widths = [
        float(cell.value) if isinstance(cell.value, (int, float, str)) else 2.0 
        for cell in sheet[4][1:]
    ]

    # FIX: Die optischen Eigenschaften via zip() direkt den Datasets übergeben
    datasets = [
        RadarDataset(
            name=name,
            values=[],
            color=color,
            marker=marker,
            line_width=width
        )
        for name, color, marker, width in zip(dataset_names, dataset_colors, dataset_markers, dataset_widths)
    ]

    # Die echten Daten starten ab Zeile 5
    for row in sheet.iter_rows(min_row=5, values_only=True):
        if row[0] is None:
            raise ValueError("Missing label in Data sheet.")

        labels.append(str(row[0]))

        for i, dataset in enumerate(datasets):
            value = row[i + 1]
            if value is None:
                raise ValueError("Found empty cell in dataset.")

            if not isinstance(value, (int, float)):
                raise TypeError(
                    f"Expected numeric value, got {type(value).__name__}: {value}"
                )

            dataset.values.append(float(value))

    return labels, datasets


def _load_style(sheet: Worksheet) -> RadarStyle:
    """Liest alle Style-Einstellungen ab Zeile 2."""
    values = {}

    for key, value in sheet.iter_rows(min_row=2, values_only=True):
        if key is None:
            continue
        values[str(key)] = value

    return RadarStyle(
        frame=values["frame"],
        line_color=values["line_color"],
        fill_color=values["fill_color"],
        fill=values["fill"],
        alpha=values["alpha"],
        ring_count=values["ring_count"],
        frame_width=values["frame_width"],
        grid_width=values["grid_width"],
        label_distance=values["label_distance"],
        scale_label_offset=values["scale_label_offset"],
        grid_alpha=values["grid_alpha"],
        dpi=int(values.get("dpi", 300))
    )


def load_chart(filepath: str) -> RadarChart:
    """Lädt ein RadarChart aus einer Excel-Datei."""
    workbook = load_workbook(filepath)

    validate_chart_excel(workbook)

    meta_sheet = workbook["Meta"]
    data_sheet = workbook["Data"]
    style_sheet = workbook["Style"]

    title = _load_meta(meta_sheet)
    labels, datasets = _load_data(data_sheet)
    style = _load_style(style_sheet)

    data = RadarData(
        title=title,
        labels=labels,
        datasets=datasets
    )

    return RadarChart(
        data=data,
        style=style
    )


def _save_meta(sheet, chart: RadarChart):
    sheet["A1"] = "Property"
    sheet["B1"] = "Value"
    sheet["A2"] = "title"
    sheet["B2"] = chart.data.title


def _save_data(sheet, chart: RadarChart):
    # Header-Meta-Informationen in Spalte A (Zeile 1 bis 4)
    sheet.cell(row=1, column=1, value="label / meta")
    sheet.cell(row=2, column=1, value="color")
    sheet.cell(row=3, column=1, value="marker")
    sheet.cell(row=4, column=1, value="line_width")

    for col, ds in enumerate(chart.data.datasets, start=2):
        sheet.cell(row=1, column=col, value=ds.name)
        sheet.cell(row=2, column=col, value=ds.color)
        sheet.cell(row=3, column=col, value=ds.marker)
        sheet.cell(row=4, column=col, value=ds.line_width)

    # FIX: Mit dem Listen-Index (idx) arbeiten, da die Excel-Zeilen (row_idx) verschoben sind
    for idx, label in enumerate(chart.data.labels):
        row_idx = idx + 5
        sheet.cell(row=row_idx, column=1, value=label)

        for col_idx, ds in enumerate(chart.data.datasets, start=2):
            sheet.cell(
                row=row_idx,
                column=col_idx,
                value=ds.values[idx]  # Nutzt jetzt sicher den korrekten Array-Index (0, 1, 2...)
            )


def _save_style(sheet, style: RadarStyle):
    sheet["A1"] = "Property"
    sheet["B1"] = "Value"
    style_dict = {
        "frame": style.frame,
        "line_color": style.line_color,
        "fill_color": style.fill_color,
        "fill": style.fill,
        "alpha": style.alpha,
        "ring_count": style.ring_count,
        "frame_width": style.frame_width,
        "grid_width": style.grid_width,
        "label_distance": style.label_distance,
        "scale_label_offset": style.scale_label_offset,
        "grid_alpha": style.grid_alpha,
        "dpi": style.dpi,
    }

    for row, (key, value) in enumerate(style_dict.items(), start=2):
        sheet.cell(row=row, column=1, value=key)
        sheet.cell(row=row, column=2, value=value)


def save_chart(chart: RadarChart, filepath: str):
    """Speichert ein RadarChart als Excel-Datei."""
    workbook = Workbook()

    if "Sheet" in workbook.sheetnames:
        del workbook["Sheet"]
    meta_sheet = workbook.create_sheet("Meta", 0)
    data_sheet = workbook.create_sheet("Data", 1)
    style_sheet = workbook.create_sheet("Style", 2)

    _save_meta(meta_sheet, chart)
    _save_data(data_sheet, chart)
    _save_style(style_sheet, chart.style)

    workbook.save(filepath)