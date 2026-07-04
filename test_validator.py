import os
from openpyxl import load_workbook
from excel_handler import load_chart

def setup_base_file():
    """Stellt sicher, dass eine valide Basisdatei aus dem Roundtrip-Test existiert."""
    base_path = "data/excel/roundtrip_test.xlsx"
    if not os.path.exists(base_path):
        print("Führe zuerst 'test_roundtrip.py' aus, um die Basisdatei zu erstellen!")
        from test_roundtrip import run_roundtrip_test
        run_roundtrip_test()
    return base_path

def run_validator_tests():
    base_filepath = setup_base_file()
    temp_filepath = "data/excel/validator_corrupt_test.xlsx"
    
    print("=== Starte Excel Validator Tests ===")

    # --- TEST 1: Fehlendes Tabellenblatt ---
    wb = load_workbook(base_filepath)
    del wb["Style"]  # Wir löschen das Style-Sheet absichtlich
    wb.save(temp_filepath)
    
    try:
        load_chart(temp_filepath)
        print("❌ Test 1 fehlgeschlagen: Fehlendes Sheet 'Style' wurde NICHT erkannt!")
    except ValueError as e:
        if "Tabellenblatt 'Style' fehlt" in str(e):
            print("✅ Test 1 erfolgreich: Fehlendes Sheet 'Style' wurde korrekt abgefangen.")
        else:
            print(f"❌ Test 1 fehlgeschlagen: Unerwarteter Fehler: {e}")

    # --- TEST 2: Leerer Titel in Meta ---
    wb = load_workbook(base_filepath)
    wb["Meta"]["B2"] = ""  # Titel löschen
    wb.save(temp_filepath)
    
    try:
        load_chart(temp_filepath)
        print("❌ Test 2 fehlgeschlagen: Leerer Titel wurde NICHT erkannt!")
    except ValueError as e:
        if "Diagrammtitel in Zelle B2 darf nicht leer sein" in str(e):
            print("✅ Test 2 erfolgreich: Leerer Titel in Meta wurde korrekt abgefangen.")
        else:
            print(f"❌ Test 2 fehlgeschlagen: Unerwarteter Fehler: {e}")

    # --- TEST 3: Fehlendes Label in Data ---
    wb = load_workbook(base_filepath)
    wb["Data"]["A2"] = None  # Das erste Label (z.B. 'Angriff') weglöschen
    wb.save(temp_filepath)
    
    try:
        load_chart(temp_filepath)
        print("❌ Test 3 fehlgeschlagen: Fehlendes Label wurde NICHT erkannt!")
    except ValueError as e:
        if "fehlt das Label in Spalte A" in str(e):
            print("✅ Test 3 erfolgreich: Fehlendes Label in Data wurde korrekt abgefangen.")
        else:
            print(f"❌ Test 3 fehlgeschlagen: Unerwarteter Fehler: {e}")

    # --- TEST 4: Falscher Datentyp (Text statt Zahl) ---
    wb = load_workbook(base_filepath)
    wb["Data"]["B2"] = "Zweiundachtzig"  # Text statt einer Zahl eintragen
    wb.save(temp_filepath)
    
    try:
        load_chart(temp_filepath)
        print("❌ Test 4 fehlgeschlagen: Text statt Zahl wurde NICHT erkannt!")
    except TypeError as e:
        if "Ungültiger Datentyp" in str(e):
            print("✅ Test 4 erfolgreich: Falscher Datentyp (Text) wurde korrekt abgefangen.")
        else:
            print(f"❌ Test 4 fehlgeschlagen: Unerwarteter Fehler: {e}")

    # --- TEST 5: Fehlender Style-Parameter ---
    wb = load_workbook(base_filepath)
    # Wir überschreiben die Zeile mit dem 'frame' Parameter in Spalte A
    for row in wb["Style"].iter_rows(min_row=2):
        if row[0].value == "frame":
            row[0].value = "kaputter_key"
            break
    wb.save(temp_filepath)
    
    try:
        load_chart(temp_filepath)
        print("❌ Test 5 fehlgeschlagen: Fehlender Style-Parameter wurde NICHT erkannt!")
    except ValueError as e:
        if "Folgende Style-Parameter fehlen: frame" in str(e):
            print("✅ Test 5 erfolgreich: Fehlender Style-Parameter wurde korrekt abgefangen.")
        else:
            print(f"❌ Test 5 fehlgeschlagen: Unerwarteter Fehler: {e}")

    # Aufräumen
    if os.path.exists(temp_filepath):
        os.remove(temp_filepath)
        
    print("\n=== Validator Tests abgeschlossen ===")

if __name__ == "__main__":
    run_validator_tests()