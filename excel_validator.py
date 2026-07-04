# excel_validator.py
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

def validate_excel_structure(workbook) -> None:
    """Prüft, ob alle notwendigen Tabellenblätter existieren."""
    required_sheets = ["Meta", "Data", "Style"]
    for sheet_name in required_sheets:
        if sheet_name not in workbook.sheetnames:
            raise ValueError(f"Ungültiges Excel-Format: Das Tabellenblatt '{sheet_name}' fehlt!")

def validate_meta_sheet(sheet: Worksheet) -> None:
    """Prüft das Meta-Tabellenblatt."""
    if sheet["B2"].value is None or str(sheet["B2"].value).strip() == "":
        raise ValueError("Validierungsfehler im Sheet 'Meta': Der Diagrammtitel in Zelle B2 darf nicht leer sein!")

def validate_data_sheet(sheet: Worksheet) -> None:
    """Prüft das Data-Tabellenblatt auf korrekte Labels, Datentypen und Längen."""
    # 1. Datensatz-Namen extrahieren (Spalte 2 bis Ende in Zeile 1)
    dataset_names = [cell.value for cell in sheet[1][1:] if cell.value is not None]
    if not dataset_names:
        raise ValueError("Validierungsfehler im Sheet 'Data': Es wurden keine Datensätze (Spaltenüberschriften) gefunden!")

    # 2. Zeilen validieren
    for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        label = row[0]
        
        # Wenn die Zeile komplett leer ist, ignorieren wir sie (oder brechen ab)
        if label is None and any(val is not None for val in row[1:]):
            raise ValueError(f"Validierungsfehler im Sheet 'Data': In Zeile {row_idx} fehlt das Label in Spalte A!")
        
        if label is None:
            continue  # Leere Zeile am Ende der Datei

        # Werte prüfen
        for col_idx, value in enumerate(row[1:len(dataset_names)+1], start=2):
            if value is None:
                raise ValueError(
                    f"Validierungsfehler im Sheet 'Data': Fehlender Wert in Zeile {row_idx} "
                    f"für den Datensatz '{sheet.cell(row=1, column=col_idx).value}'!"
                )
            if not isinstance(value, (int, float)):
                raise TypeError(
                    f"Validierungsfehler im Sheet 'Data': Ungültiger Datentyp in Zeile {row_idx}, Spalte {col_idx}. "
                    f"Erwartet: Zahl, Gefunden: '{value}' ({type(value).__name__})!"
                )

def validate_style_sheet(sheet: Worksheet) -> None:
    """Prüft, ob alle essenziellen Style-Eigenschaften vorhanden sind."""
    required_styles = [
        "frame", "line_color", "fill_color", "fill", "alpha", "ring_count",
        "frame_width", "grid_width", "label_distance", "scale_label_offset", "grid_alpha"
    ]
    
    found_keys = set()
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[0] is not None:
            found_keys.add(str(row[0]).strip())
            
    missing_keys = [style for style in required_styles if style not in found_keys]
    if missing_keys:
        raise ValueError(f"Validierungsfehler im Sheet 'Style': Folgende Style-Parameter fehlen: {', '.join(missing_keys)}!")

def validate_chart_excel(workbook) -> None:
    """Führt alle Validierungen nacheinander aus."""
    validate_excel_structure(workbook)
    validate_meta_sheet(workbook["Meta"])
    validate_data_sheet(workbook["Data"])
    validate_style_sheet(workbook["Style"])